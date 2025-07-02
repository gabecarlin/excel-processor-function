import azure.functions as func
import json
import base64
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import PatternFill, Font
import logging

def main(req: func.HttpRequest) -> func.HttpResponse:
    logging.info('Python HTTP trigger function processed a request.')
    
    try:
        # Get request data from Power Automate
        req_body = req.get_json()
        
        # Extract file info
        filename = req_body.get('filename', 'input.xlsx')
        file_content = req_body.get('content')  # base64 encoded from Power Automate
        
        if not file_content:
            return func.HttpResponse(
                json.dumps({"error": "No file content provided"}),
                status_code=400,
                mimetype="application/json"
            )
        
        # Decode the Excel file
        excel_data = base64.b64decode(file_content)
        
        # Process the Excel file
        result = process_excel_file(excel_data, filename)
        
        # Return results
        return func.HttpResponse(
            json.dumps(result),
            mimetype="application/json"
        )
        
    except Exception as e:
        logging.error(f"Error processing request: {str(e)}")
        return func.HttpResponse(
            json.dumps({"error": str(e)}),
            status_code=500,
            mimetype="application/json"
        )

def process_excel_file(file_data, filename):
    """Process Excel file and return summary statistics"""
    
    # Read Excel file from bytes
    excel_file = io.BytesIO(file_data)
    
    # Get all sheet names
    xls = pd.ExcelFile(excel_file)
    sheet_names = xls.sheet_names
    
    # Initialize results
    results = {
        "filename": filename,
        "sheets_processed": [],
        "summary": {},
        "charts_created": False
    }
    
    all_data = {}
    
    # Process each sheet
    for sheet_name in sheet_names:
        try:
            # Read the sheet
            df = pd.read_excel(excel_file, sheet_name=sheet_name)
            
            # Get numeric columns only
            numeric_cols = df.select_dtypes(include=['number']).columns.tolist()
            
            if numeric_cols:
                # Calculate statistics for numeric columns
                sheet_stats = {}
                for col in numeric_cols:
                    col_data = df[col].dropna()
                    if len(col_data) > 0:
                        sheet_stats[col] = {
                            "mean": float(col_data.mean()),
                            "median": float(col_data.median()),
                            "std": float(col_data.std()) if len(col_data) > 1 else 0,
                            "min": float(col_data.min()),
                            "max": float(col_data.max()),
                            "count": int(len(col_data))
                        }
                
                results["summary"][sheet_name] = sheet_stats
                all_data[sheet_name] = df
                results["sheets_processed"].append(sheet_name)
                
        except Exception as e:
            results["sheets_processed"].append(f"{sheet_name} (ERROR: {str(e)})")
    
    # Create output Excel file with results
    if results["summary"]:
        output_file = create_summary_excel(results["summary"], all_data)
        
        # Convert to base64 for Power Automate
        results["output_file"] = base64.b64encode(output_file).decode('utf-8')
        results["output_filename"] = f"processed_{filename}"
    
    return results

def create_summary_excel(summary_data, original_data):
    """Create Excel file with summary statistics and charts"""
    
    # Create new workbook
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create summary sheet
    summary_sheet = wb.create_sheet("Summary_Statistics")
    
    # Header styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    row = 1
    for sheet_name, stats in summary_data.items():
        # Sheet header
        summary_sheet.cell(row=row, column=1, value=f"Sheet: {sheet_name}")
        summary_sheet.cell(row=row, column=1).font = Font(bold=True, size=14)
        row += 2
        
        # Column headers
        headers = ["Column", "Mean", "Median", "Std Dev", "Min", "Max", "Count"]
        for col, header in enumerate(headers, 1):
            cell = summary_sheet.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
        row += 1
        
        # Data rows
        for col_name, col_stats in stats.items():
            summary_sheet.cell(row=row, column=1, value=col_name)
            summary_sheet.cell(row=row, column=2, value=round(col_stats["mean"], 2))
            summary_sheet.cell(row=row, column=3, value=round(col_stats["median"], 2))
            summary_sheet.cell(row=row, column=4, value=round(col_stats["std"], 2))
            summary_sheet.cell(row=row, column=5, value=round(col_stats["min"], 2))
            summary_sheet.cell(row=row, column=6, value=round(col_stats["max"], 2))
            summary_sheet.cell(row=row, column=7, value=col_stats["count"])
            row += 1
        
        row += 2  # Space between sheets
    
    # Copy original sheets with highlighting
    for sheet_name, df in original_data.items():
        ws = wb.create_sheet(f"Original_{sheet_name}")
        
        # Write headers with styling
        for col, header in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        # Write data
        for row_idx, row_data in enumerate(df.values, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Add simple chart if numeric data exists
        numeric_cols = df.select_dtypes(include=['number']).columns[:5]  # Limit to 5 columns
        if len(numeric_cols) > 0 and len(df) > 1:
            try:
                chart = BarChart()
                chart.title = f"Data Overview - {sheet_name}"
                chart.y_axis.title = "Values"
                chart.x_axis.title = "Records"
                
                # Add data for first numeric column
                data = Reference(ws, min_col=df.columns.get_loc(numeric_cols[0])+1, 
                               min_row=2, max_row=min(20, len(df)+1))  # Limit to 20 rows
                chart.add_data(data)
                
                ws.add_chart(chart, f"H2")
            except Exception as e:
                logging.warning(f"Could not create chart for {sheet_name}: {e}")
    
    # Auto-adjust column widths
    for sheet in wb.worksheets:
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    # Save to bytes
    output_buffer = io.BytesIO()
    wb.save(output_buffer)
    output_buffer.seek(0)
    
    return output_buffer.getvalue()
